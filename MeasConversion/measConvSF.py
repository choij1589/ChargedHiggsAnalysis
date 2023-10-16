import os
import argparse
import pandas as pd
import openpyxl
import ROOT
from math import pow, sqrt
from string import ascii_uppercase
from ctypes import c_double
from pprint import pprint

parser = argparse.ArgumentParser()
parser.add_argument("--era", required=True, type=str, help="era")
parser.add_argument("--channel", required=True, type=str, help="channel")
args = parser.parse_args()

histkey = "ZCand/mass"
WORKDIR = os.environ["WORKDIR"]

DATASTREAM = ""
REGION = ""
if args.channel == "Skim1E2Mu": 
    DATASTREAM = "MuonEG"
    REGION = "ZGamma1E2Mu"
if args.channel == "Skim3Mu": 
    DATASTREAM = "DoubleMuon"
    REGION = "ZGamma3Mu"
assert DATASTREAM in ["MuonEG", "DoubleMuon"]
assert REGION in ["ZGamma1E2Mu", "ZGamma3Mu"]

CONV = ["DYJets_MG", "DYJets10to50_MG", "ZGToLLG"]
DIBOSON = ["WZTo3LNu_amcatnlo","ZZTo4L_powheg"]
TTX     = ["ttWToLNu", "ttZToLLNuNu", "ttHToNonbb"]
OTHERS = ["WWW", "WWZ", "WZZ", "ZZZ", "tZq", "TTG", "tHq", "TTTT", "WWG", "VBF_HToZZTo4L", "GluGluHToZZTo4L"]
MCList = CONV + DIBOSON + TTX + OTHERS

SYSTs = []
if args.channel == "Skim1E2Mu":
    SYSTs = [("NonpromptUp", "NonpromptDown"),
             ("L1PrefireUp", "L1PrefireDown"),
             ("PileupReweightUp", "PileupReweightDown"),
             ("MuonIDSFUp", "MuonIDSFDown"),
             ("ElectronIDSFUp", "ElectronIDSFDown"),
             ("EMuTrigSFUp", "EMuTrigSFDown"),
             ("JetResUp", "JetResDown"),
             ("JetEnUp", "JetEnDown"),
             ("ElectronResUp", "ElectronResDown"),
             ("ElectronEnUp", "ElectronEnDown"),
             ("MuonEnUp", "MuonEnDown")]
             #("HeavyTagUpCorr", "HeavyTagDownCorr"),
             #("HeavyTagUpUnCorr", "HeavyTagDownUnCorr"),
             #("LightTagUpCorr", "LightTagDownCorr"),
             #("LightTagUpUnCorr", "LightTagDownUnCorr")]
if args.channel == "Skim3Mu":
    SYSTs = [("NonpromptUp", "NonpromptDown"),
             ("L1PrefireUp", "L1PrefireDown"),
             ("PileupReweightUp", "PileupReweightDown"),
             ("MuonIDSFUp", "MuonIDSFDown"),
             ("DblMuTrigSFUp", "DblMuTrigSFDown"),
             ("JetResUp", "JetResDown"),
             ("JetEnUp", "JetEnDown"),
             ("ElectronResUp", "ElectronResDown"),
             ("ElectronEnUp", "ElectronEnDown"),
             ("MuonEnUp", "MuonEnDown")]
             #("HeavyTagUpCorr", "HeavyTagDownCorr"),
             #("HeavyTagUpUnCorr", "HeavyTagDownUnCorr"),
             #("LightTagUpCorr", "LightTagDownCorr"),
             #("LightTagUpUnCorr", "LightTagDownUnCorr")]
             
## make a table
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Rate"

#### make a index column for systematics
sheet["A1"] = "process"
sheet["A2"] = "Central"
sheet["A3"] = "Stat"
for i, (systUp, systDown) in enumerate(SYSTs, start=2):
    sheet[f"A{i*2}"] = systUp
    sheet[f"A{i*2+1}"] = systDown
totalIdx = (len(SYSTs)+2)*2 
sheet[f"A{totalIdx}"] = "Total"

#### Fill out estimates rates
## data
sheet["B1"] = "DATA"
f = ROOT.TFile.Open(f"{WORKDIR}/data/MeasConversion/{args.era}/{args.channel}__/DATA/MeasConversion_SkimTree_SS2lOR3l_{DATASTREAM}.root")
h = f.Get(f"{REGION}/Central/{histkey}"); h.SetDirectory(0)
f.Close()

stat = c_double()
rate = h.IntegralAndError(0, h.GetNbinsX()+1, stat)
sheet["B2"] = rate
sheet["B3"] = stat.value

## nonprompt
f = ROOT.TFile.Open(f"{WORKDIR}/data/MeasConvMatrix/{args.era}/{args.channel}__/DATA/MeasConvMatrix_SkimTree_SS2lOR3l_{DATASTREAM}.root")
h = f.Get(f"{REGION}/Central/{histkey}"); h.SetDirectory(0)
f.Close()

stat = c_double()
rate = h.IntegralAndError(0, h.GetNbinsX()+1, stat)
sheet["C1"] = "nonprompt"
sheet["C2"] = rate
sheet["C3"] = stat.value
sheet["C4"] = rate*1.3      # NonpromptUp
sheet["C5"] = rate*0.7      # NonpromptDown

f = ROOT.TFile.Open(f"{WORKDIR}/data/MeasConvMatrix/{args.era}/{args.channel}__/MeasConvMatrix_SkimTree_SS2lOR3l_DYJets_MG.root")
h = f.Get(f"{REGION}/Central/{histkey}"); h.SetDirectory(0)
f.Close()

stat = c_double()
rate = h.IntegralAndError(0, h.GetNbinsX()+1, stat)
sheet["D1"] = "nonprompt_conv"
sheet["D2"] = rate
sheet["D3"] = stat.value
sheet["D4"] = rate*1.3      # NonpromptUp
sheet["D5"] = rate*0.7      # NonpromptDown

for cidx, sample in enumerate(MCList, start=4):
    prefix = ascii_uppercase[cidx]
    sheet[f"{prefix}1"] = sample
    fkey = f"{WORKDIR}/data/MeasConversion/{args.era}/{args.channel}__/MeasConversion_SkimTree_SS2lOR3l_{sample}.root"
    assert os.path.exists(fkey)
    f = ROOT.TFile.Open(fkey)
    try:
        h = f.Get(f"{REGION}/Central/{histkey}"); h.SetDirectory(0)
    except:
        print(sample); continue
    
    stat = c_double()
    rate = h.IntegralAndError(0, h.GetNbinsX()+1, stat)
    sheet[f"{prefix}2"] = rate
    sheet[f"{prefix}3"] = stat.value
    # skip 4 and 5 - Nonprompt
    
    for idx, (systUp, systDown) in enumerate(SYSTs[1:], start=3):
        h_up = f.Get(f"{REGION}/{systUp}/{histkey}"); h_up.SetDirectory(0)
        h_down = f.Get(f"{REGION}/{systDown}/{histkey}"); h_down.SetDirectory(0)
        sheet[f"{prefix}{idx*2}"] = h_up.Integral()
        sheet[f"{prefix}{idx*2+1}"] = h_down.Integral()
    f.Close()

for cidx, col in enumerate(sheet.iter_cols(min_row=2, min_col=2, values_only=True), start=1):
    rate = col[0]
    stat = col[1]
    if rate is None: continue
         
    total_unc = pow(stat, 2)
    for i in range(len(SYSTs)):
        rate_up, rate_down = col[(i+1)*2:(i+2)*2]
        if rate_up is None: continue 
        syst_up = abs(rate - rate_up)
        syst_down = abs(rate - rate_down)
        total_unc += pow(max(syst_up, syst_down), 2)
    total_unc = sqrt(total_unc)
    prefix = ascii_uppercase[cidx]
    sheet[f"{prefix}{totalIdx}"] = total_unc

data = sheet.values
cols = next(data)
df = pd.DataFrame(data, columns=cols)
df.set_index("process", inplace=True)
#pprint(df)

#### Measure conversion scale factors
conv_sheet = workbook.create_sheet("Measurement")
conv_sheet["A1"] = "process"
conv_sheet["B1"] = "DATA"
conv_sheet["C1"] = "nonprompt"
#conv_sheet["D1"] = "nonprompt_conv"
conv_sheet["D1"] = "conversion"
conv_sheet["E1"] = "prompt"
conv_sheet["F1"] = "conversion SF"

## Central
conv_sheet["A2"] = "Central"
conv_sheet["B2"] = df.loc["Central", "DATA"]
conv_sheet["C2"] = max(df.loc["Central", "nonprompt"] - df.loc["Central", "nonprompt_conv"], 0.)
#conv_sheet["D2"] = df.loc["Central", "nonprompt_conv"]
conv_sheet["D2"] = df.loc["Central", "DYJets_MG"]
prompt_rate = 0.
for prompt in DIBOSON + TTX + OTHERS:
    if df.loc["Central", prompt] is None: continue
    prompt_rate += df.loc["Central", prompt]
conv_sheet["E2"] = prompt_rate
conv_sheet["F2"] = (conv_sheet["B2"].value - conv_sheet["C2"].value - conv_sheet["E2"].value) / conv_sheet["D2"].value

## Stat
conv_sheet["A3"] = "Stat"
conv_sheet["B3"] = df.loc["Stat", "DATA"]
conv_sheet["C3"] = sqrt(pow(df.loc["Stat", "nonprompt"], 2)+pow(df.loc["Stat", "nonprompt_conv"], 2))
conv_sheet["D3"] = df.loc["Stat", "DYJets_MG"]
prompt_stat = 0.
for prompt in DIBOSON + TTX + OTHERS:
    if df.loc["Stat", prompt] is None: continue
    prompt_stat += pow(df.loc["Stat", prompt], 2)
conv_sheet["E3"] = sqrt(prompt_stat)

Ndata = conv_sheet["B2"].value
Nfake = conv_sheet["C2"].value
Nconv = conv_sheet["D2"].value
Npred = conv_sheet["E2"].value
dNconv = -((Ndata - Nfake - Npred) / pow(Nconv, 2))*conv_sheet["D3"].value
dNdata = (1 / Nconv)*conv_sheet["B3"].value
dNfake = (-1 / Nconv)*conv_sheet["C3"].value
dNpred = (-1 / Nconv)*conv_sheet["E3"].value
conv_sheet["F3"] = dNconv + dNdata + dNfake + dNpred

for idx, (systUp, systDown) in enumerate(SYSTs, start=2):
    conv_sheet[f"A{idx*2}"] = systUp
    conv_sheet[f"A{idx*2+1}"] = systDown
    
    conv_sheet[f"B{idx*2}"] = df.loc["Central", "DATA"]
    conv_sheet[f"B{idx*2+1}"] = df.loc["Central", "DATA"]
    
    if "Nonprompt" in systUp:
        conv_sheet[f"C{idx*2}"] = max(df.loc[systUp, "nonprompt"]-df.loc[systUp, "nonprompt_conv"], 0.)
        conv_sheet[f"C{idx*2+1}"] = max(df.loc[systDown, "nonprompt"]-df.loc[systDown, "nonprompt_conv"], 0.) 
        conv_sheet[f"D{idx*2}"] = df.loc["Central", "DYJets_MG"]
        conv_sheet[f"D{idx*2+1}"] = df.loc["Central", "DYJets_MG"]
        prompt_rate = 0.
        for prompt in DIBOSON + TTX + OTHERS: 
            if df.loc["Central", prompt] is None: continue
            prompt_rate += df.loc["Central", prompt] 
        conv_sheet[f"E{idx*2}"] = prompt_rate
        conv_sheet[f"E{idx*2+1}"] = prompt_rate
    else:
        conv_sheet[f"C{idx*2}"] = max(df.loc["Central", "nonprompt"]-df.loc["Central", "nonprompt_conv"], 0.)
        conv_sheet[f"C{idx*2+1}"] = max(df.loc["Central", "nonprompt"]- df.loc["Central", "nonprompt_conv"], 0.)
        conv_sheet[f"D{idx*2}"] = df.loc[systUp, "DYJets_MG"]
        conv_sheet[f"D{idx*2+1}"] = df.loc[systDown, "DYJets_MG"]
        prompt_rate_up = 0.
        prompt_rate_down = 0.
        for prompt in DIBOSON + TTX + OTHERS:
            if df.loc[systUp, prompt] is None: continue 
            prompt_rate_up += df.loc[systUp, prompt]
        for prompt in DIBOSON + TTX + OTHERS:
            if df.loc[systDown, prompt] is None: continue 
            prompt_rate_down += df.loc[systDown, prompt] 
        conv_sheet[f"E{idx*2}"] = prompt_rate_up
        conv_sheet[f"E{idx*2+1}"] = prompt_rate_down
    conv_sheet[f"F{idx*2}"] = (conv_sheet[f"B{idx*2}"].value - conv_sheet[f"C{idx*2}"].value - conv_sheet[f"E{idx*2}"].value) / conv_sheet[f"D{idx*2}"].value
    conv_sheet[f"F{idx*2+1}"] = (conv_sheet[f"B{idx*2+1}"].value - conv_sheet[f"C{idx*2+1}"].value - conv_sheet[f"E{idx*2+1}"].value) / conv_sheet[f"D{idx*2+1}"].value

## Measure total uncertainty
conv_sheet[f"A{totalIdx}"] = "Total"

for col in conv_sheet.iter_cols(min_row=2, min_col=6, max_col=6, values_only=True):
    sf = col[0]
    total_unc = pow(col[1], 2)
    for idx, (systUp, systDown) in enumerate(SYSTs, start=1):
        sf_up = col[idx*2] 
        sf_down = col[idx*2+1]
        #print(sf, sf_up, sf_down)
        total_unc += pow(max(abs(sf-sf_up), abs(sf-sf_down)), 2)
    total_unc = sqrt(total_unc)
conv_sheet[f"F{totalIdx}"] = total_unc

save_path = f"results/{args.era}/{args.channel}.xlsx"
os.makedirs(os.path.dirname(save_path), exist_ok=True) 
workbook.save(filename=save_path)
    


